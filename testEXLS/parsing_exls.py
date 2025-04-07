# -*- coding: cp1251 -*-
from shlex import join
import openpyxl
import os
import win 


def path_p(nam_f):
	path=os.path.join('testEXLS',nam_f)
	file_n_path=os.path.abspath(path).replace('\\','\\\\')
	return file_n_path
#D:\gitClone\KhamzinRadik\python-2.5\testEXLS\2025_otchet.xlsx


def ub_pr(strf):# убираем лишние пробелы и приводим к одному регистру
	
	strf=strf.split()
	strf=' '.join(strf)
	
	strf=strf.split()
	return strf
def poluchenie_slovarey(f_salido,f_rashod):
	file_name='2025_otchet.xlsx'
	file_name_path=path_p(file_name)

	wb = openpyxl.load_workbook(file_name_path, data_only=True)
	type(wb)


	poisk='Сальдо на конец  периода (экономия +), (перерасход -) от начисленного'
	poisk=ub_pr(poisk)
	poisk_raschod='ИТОГО стоимость услуг в год с НДС'
	poisk_raschod=ub_pr(poisk_raschod)

	sheets = wb.sheetnames#запрос названий всех листов
	sheet=wb.active#выбор активного листа

	sheets_name=[]
	for sheet in sheets:#передача названий листов в список
		sheets_name.append(sheet)
	# if 'New Sheet Result' in sheets_name:
	#     print('list est!')
	# else:
	# 	wb.create_sheet(index=0,title="New Sheet Result")
	# 	wb.save(file_name_path)
	# 	print('list sozdan!')

	#print('sheets_name',sheets_name)

	for i in range (len(sheets_name)):

		sheet = wb[sheets_name[i]]#обращаемся к листу по названию листа
		print('\n\n',sheet)
		print(sheet['A5'].value)
		dimensions = sheet.max_row

		

		for i_col in range (1,	dimensions):
		# # печатаем значение ячейки A1
				a='A'+str(i_col)
				b='B'+str(i_col)
				#print(a)
				sh=sheet[a].value
				sh=str(sh)
				sh=ub_pr(sh)
				count_raschod=0
				count=0
				for i_el in sh:
					if i_el in poisk_raschod:
						count_raschod+=1
					if i_el in poisk:
					
						count+=1
						#print(count)
				if count_raschod>6:
					#print('\n',sheet[a].value,end=': ')
					#print(sheet[b].value)
					rashod={sheet[a].value:(sheet[b].value)}
					f_rashod[str(sheet['A5'].value)]=rashod
				
				if count>9 :
					#print('\n',sheet[a].value,end=': ')
					#print(sheet[b].value)
					salido={sheet[a].value:(sheet[b].value)}
					f_salido[str(sheet['A5'].value)]=salido
	print('f_rashod',f_rashod)
	print('f_salido',f_salido)
	print('\n\n\n')




def zapis_v_excel(f_salido,f_rashod):
	file_name='2025_otchet.xlsx'
	file_name_path=path_p(file_name)

	wb = openpyxl.Workbook() 
	sheet = wb.active
	c1= sheet.cell(row = 1, column = 1)
	c1.value='улица'
	c2= sheet.cell(row = 1, column = 2)
	c2.value='сальдо'
	c3= sheet.cell(row = 1, column = 3)
	c3.value='расход'
	print(f_salido)
	i=2
	for i_el in f_salido:
		print (i)
		c1 = sheet.cell(row = i, column = 1)
		c1.value=i_el
		i+=1
	i=2	
	for k in f_salido.values():
			print('k',k)
			for key,value in k.items():
				print('key',key)
				print('value',value)
				c2 = sheet.cell(row = i, column = 2)
				c2.value=value
				i+=1
	i=2
	for k in f_rashod.values():
			print('k',k)
			for key,value in k.items():
				print('key',key)
				print('value',value)
				c3 = sheet.cell(row = i, column = 3)
				c3.value=value
				i+=1
				# c1 = sheet.cell(row = i, column = 1)
				
				# c1.value=k
		

	file_name='2025_otchet_result.xlsx'
	file_name_path=path_p(file_name)
	sheet.column_dimensions["A"].width = 40
	sheet.column_dimensions["B"].width = 15
	sheet.column_dimensions["C"].width = 15
	wb.save(file_name_path)


win.Window()




f_salido={}
f_rashod={}
poluchenie_slovarey(f_salido,f_rashod)
zapis_v_excel(f_salido,f_rashod)

# print('f_salido\n',f_salido)
# print('\n\n\n')
# print('f_rashod\n',f_rashod)
# speak_file=open(file_name_path,'r')
# file=speak_file.readlines()

# for i_str in range (len(file)-1,0,-1):

	
#      print(file[i_str])